"""
Módulo para exportar solicitudes a Excel con formato correcto.
Las fechas y horas se exportan en zona horaria de Chile.
"""

from django.http import HttpResponse
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pytz


def exportar_solicitudes_excel(solicitudes, nombre_archivo='solicitudes'):
    """
    Exporta solicitudes a Excel con formato profesional.
    
    IMPORTANTE: Las fechas/horas ya vienen convertidas a hora de Chile
    automáticamente por Django, no necesitas hacer conversión manual.
    
    Args:
        solicitudes: QuerySet de Solicitud
        nombre_archivo: Nombre base del archivo (sin extensión)
    
    Returns:
        HttpResponse con el archivo Excel
    """
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Solicitudes"
    
    # Estilos
    header_fill = PatternFill(start_color="00B4D8", end_color="00B4D8", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = [
        'ID', 'Fecha', 'Hora', 'Tipo', 'N° Pedido/ST', 'Cliente',
        'Códigos', 'Estado', 'Urgente', 'Transporte', 'Bodega',
        'Observación', 'Solicitante', 'Creado el', 'Actualizado el'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Datos
    chile_tz = pytz.timezone('America/Santiago')
    
    for row_num, solicitud in enumerate(solicitudes, 2):
        # NOTA: solicitud.fecha_solicitud ya está en hora de Chile gracias a Django
        # NOTA: solicitud.created_at es timezone-aware, lo convertimos a Chile
        
        # Convertir created_at y updated_at a hora de Chile
        created_chile = solicitud.created_at.astimezone(chile_tz)
        updated_chile = solicitud.updated_at.astimezone(chile_tz)
        
        # Referencia (Pedido o ST)
        referencia = ''
        if solicitud.tipo == 'ST' and solicitud.numero_st:
            referencia = solicitud.numero_st
        elif solicitud.numero_pedido:
            referencia = solicitud.numero_pedido
        
        # Contar códigos
        total_codigos = solicitud.detalles.count() or (1 if solicitud.codigo else 0)
        
        # Datos de la fila
        row_data = [
            solicitud.id,
            solicitud.fecha_solicitud,  # Ya está en Chile ✅
            solicitud.hora_solicitud,   # Ya está en Chile ✅
            solicitud.get_tipo_display(),
            referencia,
            solicitud.cliente,
            total_codigos,
            solicitud.get_estado_display(),
            'Sí' if solicitud.urgente else 'No',
            solicitud.get_transporte_display(),
            solicitud.bodega or '-',
            solicitud.observacion or '-',
            solicitud.solicitante.username if solicitud.solicitante else 'Sistema',
            created_chile.strftime('%d/%m/%Y %H:%M:%S'),  # Convertido a Chile ✅
            updated_chile.strftime('%d/%m/%Y %H:%M:%S'),  # Convertido a Chile ✅
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Ajustar anchos de columna
    for col_num in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 15
    
    # Preparar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Nombre del archivo con fecha actual de Chile
    fecha_actual = timezone.now().astimezone(chile_tz).strftime('%Y%m%d_%H%M')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}_{fecha_actual}.xlsx"'
    
    wb.save(response)
    return response


def exportar_kpis_excel(fecha_inicio, fecha_fin):
    """
    Exporta KPIs de solicitudes a Excel.
    
    IMPORTANTE: Todas las fechas/horas ya están en zona horaria de Chile.
    Django maneja automáticamente la conversión de UTC a Chile.
    
    Args:
        fecha_inicio: date object (ya en hora Chile)
        fecha_fin: date object (ya en hora Chile)
    
    Returns:
        HttpResponse con el archivo Excel
    """
    from solicitudes.models import Solicitud
    from django.db.models import Count, Avg, F, ExpressionWrapper, DurationField
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "KPIs"
    
    # Filtrar solicitudes por rango de fechas
    # NOTA: fecha_solicitud ya está en hora de Chile, no necesitas conversión
    solicitudes = Solicitud.objects.filter(
        fecha_solicitud__range=[fecha_inicio, fecha_fin]
    )
    
    # Calcular KPIs
    total_solicitudes = solicitudes.count()
    urgentes = solicitudes.filter(urgente=True).count()
    por_estado = solicitudes.values('estado').annotate(total=Count('id'))
    por_tipo = solicitudes.values('tipo').annotate(total=Count('id'))
    
    # Escribir KPIs en Excel
    ws['A1'] = 'KPI'
    ws['B1'] = 'Valor'
    ws['A1'].font = Font(bold=True)
    ws['B1'].font = Font(bold=True)
    
    row = 2
    ws[f'A{row}'] = 'Período'
    ws[f'B{row}'] = f"{fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
    
    row += 1
    ws[f'A{row}'] = 'Total Solicitudes'
    ws[f'B{row}'] = total_solicitudes
    
    row += 1
    ws[f'A{row}'] = 'Solicitudes Urgentes'
    ws[f'B{row}'] = urgentes
    
    row += 2
    ws[f'A{row}'] = 'Por Estado:'
    ws[f'A{row}'].font = Font(bold=True)
    
    for estado_data in por_estado:
        row += 1
        ws[f'A{row}'] = f"  {estado_data['estado']}"
        ws[f'B{row}'] = estado_data['total']
    
    row += 2
    ws[f'A{row}'] = 'Por Tipo:'
    ws[f'A{row}'].font = Font(bold=True)
    
    for tipo_data in por_tipo:
        row += 1
        ws[f'A{row}'] = f"  {tipo_data['tipo']}"
        ws[f'B{row}'] = tipo_data['total']
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    chile_tz = pytz.timezone('America/Santiago')
    fecha_actual = timezone.now().astimezone(chile_tz).strftime('%Y%m%d_%H%M')
    response['Content-Disposition'] = f'attachment; filename="kpis_{fecha_actual}.xlsx"'
    
    wb.save(response)
    return response

